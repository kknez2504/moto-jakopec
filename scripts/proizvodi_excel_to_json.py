"""
Pretvara proizvodi.xlsx u:
  - src/data/products.json
  - src/data/brands.json

Pokreni nakon svake izmjene u Excelu:
    py -X utf8 scripts/proizvodi_excel_to_json.py
"""

import json, os, re
import openpyxl

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR    = os.path.join(SCRIPT_DIR, "..")
EXCEL_PATH  = os.path.join(ROOT_DIR, "proizvodi.xlsx")
JSON_PATH   = os.path.join(ROOT_DIR, "src", "data", "products.json")
BRANDS_PATH = os.path.join(ROOT_DIR, "src", "data", "brands.json")
IMG_BASE    = os.path.join(ROOT_DIR, "public", "images")

IMG_EXTS = {".jpg", ".jpeg", ".png", ".webp"}

def natural_sort_key(s):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]

def get_images(image_folder: str, category: str) -> list:
    if not image_folder:
        return []
    sub = category.lower() if category.lower() in ("oprema", "dijelovi") else "oprema"
    folder_path = os.path.join(IMG_BASE, sub, image_folder)
    if not os.path.isdir(folder_path):
        return []
    files = [f for f in os.listdir(folder_path)
             if os.path.splitext(f)[1].lower() in IMG_EXTS]
    files.sort(key=natural_sort_key)
    return [f"/images/{sub}/{image_folder}/{f}" for f in files]

def parse_bool(val) -> bool:
    if isinstance(val, bool): return val
    if isinstance(val, str):  return val.strip().upper() in ("TRUE", "DA", "1", "YES")
    if isinstance(val, (int, float)): return bool(val)
    return False

def parse_val(val):
    if val is None: return None
    if isinstance(val, str):
        s = val.strip()
        return s if s else None
    return val

# ── Čitanje brendova iz sheeta ───────────────────────────────────────────────
def read_brands(wb, sheet_name, category_key):
    """Čita sheet brendova i vraća listu dict objekata."""
    if sheet_name not in wb.sheetnames:
        print(f"  UPOZORENJE: Sheet '{sheet_name}' ne postoji — preskačem brendove za {category_key}")
        return []

    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    brands  = []

    for row in rows[1:]:
        d = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}

        brand_id    = parse_val(d.get("brand_id"))
        name        = parse_val(d.get("name"))
        logo_folder = parse_val(d.get("logo_folder"))
        website     = parse_val(d.get("website"))
        active      = parse_bool(d.get("active", True))

        if not brand_id or not name:
            continue

        brands.append({
            "brand_id": brand_id,
            "name":     name,
            "logo":     f"/images/brendovi/{logo_folder}.svg" if logo_folder else "",
            "website":  website or "",
            "active":   active,
        })

    return brands

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"GREŠKA: {EXCEL_PATH} ne postoji.")
        print("Pokreni prvo: py -X utf8 scripts/kreiraj_excel_predlozak.py")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # ── 1. Brendovi ──────────────────────────────────────────────────────────
    oprema_brands   = read_brands(wb, "Brendovi_Oprema",   "oprema")
    dijelovi_brands = read_brands(wb, "Brendovi_Dijelovi", "dijelovi")

    brands_data = {
        "oprema":   oprema_brands,
        "dijelovi": dijelovi_brands,
    }

    os.makedirs(os.path.dirname(BRANDS_PATH), exist_ok=True)
    with open(BRANDS_PATH, "w", encoding="utf-8") as f:
        json.dump(brands_data, f, ensure_ascii=False, indent=2)

    print(f"Brendovi → {BRANDS_PATH}")
    print(f"  Oprema:   {len(oprema_brands)}")
    print(f"  Dijelovi: {len(dijelovi_brands)}")

    # ── 2. Proizvodi ─────────────────────────────────────────────────────────
    if "Proizvodi" not in wb.sheetnames:
        print("GREŠKA: Sheet 'Proizvodi' ne postoji!")
        return

    ws   = wb["Proizvodi"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("GREŠKA: Excel je prazan!")
        return

    headers  = [str(h).strip() if h else "" for h in rows[0]]
    products = []
    skipped  = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        d = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}

        name = parse_val(d.get("name"))
        if not name:
            skipped += 1
            continue

        raw_id = d.get("id")
        try:
            pid = int(raw_id) if raw_id is not None else row_idx - 1
        except (ValueError, TypeError):
            pid = row_idx - 1

        # category MORA biti pročitan PRIJE get_images()
        category     = parse_val(d.get("category")) or ""
        subcategory  = parse_val(d.get("subcategory")) or ""
        image_folder = parse_val(d.get("image_folder")) or ""
        brand        = parse_val(d.get("brand")) or ""
        images       = get_images(image_folder, category)

        # Validacija kategorije
        if category not in ("Oprema", "Dijelovi"):
            print(f"  UPOZORENJE red {row_idx}: category='{category}' nije 'Oprema' ni 'Dijelovi' — preskačem")
            skipped += 1
            continue

        product = {
            "id":           pid,
            "name":         name,
            "category":     category,
            "subcategory":  subcategory,
            "price":        parse_val(d.get("price")) or "",
            "description":  parse_val(d.get("description")) or "",
            "image_folder": image_folder,
            "brand":        brand,
            "is_new":       parse_bool(d.get("is_new", False)),
            "images":       images,
        }
        products.append(product)

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    oprema   = sum(1 for p in products if p["category"] == "Oprema")
    dijelovi = sum(1 for p in products if p["category"] == "Dijelovi")

    print(f"\nProizvodi → {JSON_PATH}")
    print(f"  Oprema:     {oprema}")
    print(f"  Dijelovi:   {dijelovi}")
    print(f"  Ukupno:     {len(products)}")
    if skipped:
        print(f"  Preskočeno: {skipped}")
    print()
    print("Sljedeći koraci:")
    print("  git add src/data/products.json src/data/brands.json public/images/")
    print("  git commit -m 'Ažurirani proizvodi i brendovi'")
    print("  git push")


if __name__ == "__main__":
    main()
