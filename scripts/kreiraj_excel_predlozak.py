"""
Postavlja Excel strukturu za upravljanje opremom i dijelovima.
Kreira/ažurira sheetove: Proizvodi, Drop Down, Brendovi_Oprema,
Brendovi_Dijelovi, Upute.

Pokreni JEDNOM za inicijalizaciju, i ponovo kad dodaješ brendove:
    py -X utf8 scripts/kreiraj_excel_predlozak.py

NAPOMENA o padajućim izbornicima:
  Podkategorije i brendovi na sheetu "Drop Down" koriste OFFSET formulu —
  dovoljno je ručno dodati novi red u "Drop Down" i dropdown se odmah ažurira
  BEZ ponovnog pokretanja ove skripte.
  Za nove BRENDOVE ipak pokreni skriptu jer treba ažurirati i brands.json.
"""

import os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR    = os.path.join(SCRIPT_DIR, "..")
EXCEL_PATH  = os.path.join(ROOT_DIR, "proizvodi.xlsx")
JSON_PATH   = os.path.join(ROOT_DIR, "src", "data", "products.json")
DD_SHEET    = "Drop Down"   # naziv sheeta s listama za padajuće izbornike

# ── Podaci brendova ──────────────────────────────────────────────────────────
BRENDOVI_OPREMA = [
    ("wind",      "WIND Raceware", "wind",      "wind-raceware.com",  True),
    ("airoh",     "Airoh",         "airoh",      "airoh.com",          True),
    ("progrip",   "Progrip",       "progrip",    "progrip.it",         True),
    ("mitas",     "Mitas",         "mitas",      "mitas-moto.com",     True),
    ("motul",     "Motul",         "motul",      "motul.com",          True),
    ("valvoline", "Valvoline",     "valvoline",  "valvoline.com",      True),
]

BRENDOVI_DIJELOVI = [
    ("prox",   "Pro-X Racing Parts", "prox",   "pro-x.com",          True),
    ("wiseco", "Wiseco",             "wiseco",  "wiseco.com",         True),
    ("wrp",    "WRP",                "wrp",     "wrp-racing.com",     True),
    ("cht",    "CHT Sprockets",      "cht",     "cht-sprockets.com",  True),
    ("rk",     "RK Chains",          "rk",      "rk-chain.com",       True),
    ("trw",    "TRW Lucas",          "trw",     "trwaftermarket.com", True),
    ("denso",  "Denso",              "denso",   "denso.com",          True),
]

KATEGORIJE = ["Oprema", "Dijelovi"]

# Odvojene liste — korisnik dodaje ispod zadnjeg reda u odgovarajući stupac
PODKATEGORIJE_OPREMA = [
    "Kacige", "Jakne", "Hlače", "Rukavice", "Čizme",
    "Odijela", "Zaštita", "Prsluk", "Termo rublje", "Ostalo",
]

PODKATEGORIJE_DIJELOVI = [
    "Filteri", "Ulja i maziva", "Lanci", "Zupčanici", "Gume",
    "Kočnice", "Svjećice", "Amortizeri", "Ovjesi", "Električni dijelovi",
    "Motor", "Ostalo",
]

BRAND_HEADERS = ["brand_id", "name", "logo_folder", "website", "active"]

# ── Stilovi ──────────────────────────────────────────────────────────────────
def header_style(color="1B5E20"):
    return (
        PatternFill("solid", fgColor=color),
        Font(color="FFFFFF", bold=True, size=10),
        Alignment(horizontal="center", vertical="center"),
        Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        ),
    )

def data_border():
    return Border(
        left=Side(style="thin", color="EEEEEE"),
        right=Side(style="thin", color="EEEEEE"),
        bottom=Side(style="thin", color="EEEEEE"),
    )

FILLS = [
    PatternFill("solid", fgColor="F9FBF9"),
    PatternFill("solid", fgColor="E8F5E9"),
]

# ── "Drop Down" sheet ─────────────────────────────────────────────────────────
def write_dropdown_sheet(wb):
    """
    Kreira sheet 'Drop Down' sa 4 stupca:
      A: Kategorija   B: Podkategorija   C: Brand Oprema   D: Brand Dijelovi

    Padajući izbornici u Proizvodi sheetu koriste OFFSET+COUNTA dinamičke
    named rangeve koji se automatski proširuju kad korisnik doda novi red.
    """
    if DD_SHEET in wb.sheetnames:
        del wb[DD_SHEET]
    ws = wb.create_sheet(DD_SHEET)

    cols = {
        "A": ("Kategorija",           KATEGORIJE,                        "1B5E20"),
        "B": ("Podkat - Oprema",      PODKATEGORIJE_OPREMA,              "E65100"),
        "C": ("Podkat - Dijelovi",    PODKATEGORIJE_DIJELOVI,            "1565C0"),
        "D": ("Brand Oprema",         [b[0] for b in BRENDOVI_OPREMA],   "7B1FA2"),
        "E": ("Brand Dijelovi",       [b[0] for b in BRENDOVI_DIJELOVI], "00838F"),
    }

    db = data_border()
    for col_letter, (title, values, color) in cols.items():
        fill, font, center, border = header_style(color)
        cell = ws[f"{col_letter}1"]
        cell.value = title
        cell.font = font; cell.fill = fill
        cell.alignment = center; cell.border = border
        ws.row_dimensions[1].height = 26
        ws.column_dimensions[col_letter].width = 24

        for ri, val in enumerate(values, start=2):
            c = ws[f"{col_letter}{ri}"]
            c.value = val
            c.fill = FILLS[ri % 2]
            c.alignment = Alignment(vertical="center")
            c.border = db
            ws.row_dimensions[ri].height = 17

    ws.freeze_panes = "A2"

    # ── Komentar za korisnike ─────────────────────────────────────────────────
    note_row = max(len(KATEGORIJE), len(PODKATEGORIJE_OPREMA),
                   len(PODKATEGORIJE_DIJELOVI),
                   len(BRENDOVI_OPREMA), len(BRENDOVI_DIJELOVI)) + 4
    ws.cell(row=note_row, column=1,
            value="Dodaj nove vrijednosti ISPOD zadnjeg reda — automatski ce se pojaviti u dropdownu.").font = \
        Font(italic=True, color="888888", size=9)
    ws.merge_cells(f"A{note_row}:D{note_row}")

    print(f"  Sheet '{DD_SHEET}': {len(KATEGORIJE)} kat, "
          f"{len(PODKATEGORIJE_OPREMA)} opr.podkat, "
          f"{len(PODKATEGORIJE_DIJELOVI)} dij.podkat, "
          f"{len(BRENDOVI_OPREMA)} br.opr, {len(BRENDOVI_DIJELOVI)} br.dij")

    # ── Dinamički named rangevi (OFFSET + COUNTA) ─────────────────────────────
    #   OFFSET(header_cell, 1 red dolje, 0 kol desno,
    #          COUNTA(cijeli stupac)-1 za header, 1 stupac)
    #   Kad korisnik doda novi red, COUNTA se automatski uvećava.
    q = f"'{DD_SHEET}'"   # ime sheeta s navodnicima (ima razmak)

    dynamic_ranges = {
        # Kategorija
        "NR_Kategorija":    f"OFFSET({q}!$A$1,1,0,COUNTA({q}!$A:$A)-1,1)",
        # Podkategorije — imenuju se točno "Oprema" i "Dijelovi" da INDIRECT radi:
        #   subcategory formula = INDIRECT($C2) → kad C2="Oprema" → traži NR "Oprema"
        "Oprema":           f"OFFSET({q}!$B$1,1,0,COUNTA({q}!$B:$B)-1,1)",
        "Dijelovi":         f"OFFSET({q}!$C$1,1,0,COUNTA({q}!$C:$C)-1,1)",
        # Brendovi — INDIRECT("NR_Brand_"&$C2) → "NR_Brand_Oprema" ili "NR_Brand_Dijelovi"
        "NR_Brand_Oprema":  f"OFFSET({q}!$D$1,1,0,COUNTA({q}!$D:$D)-1,1)",
        "NR_Brand_Dijelovi":f"OFFSET({q}!$E$1,1,0,COUNTA({q}!$E:$E)-1,1)",
    }

    for name in list(dynamic_ranges.keys()):
        if name in wb.defined_names:
            del wb.defined_names[name]
    for name, formula in dynamic_ranges.items():
        wb.defined_names[name] = DefinedName(name, attr_text=formula)

    return dynamic_ranges

# ── Padajući izbornici u Proizvodi sheetu ─────────────────────────────────────
def add_dropdowns(ws):
    MAX_ROW = 10000

    col_map = {}
    for ci in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=ci).value
        if val:
            col_map[str(val).strip()] = get_column_letter(ci)

    def make_dv(formula, prompt_title, prompt_body, strict=True):
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showInputMessage=True,
            promptTitle=prompt_title,
            prompt=prompt_body,
            showErrorMessage=strict,
            errorTitle="Neispravna vrijednost" if strict else "",
            error='Mora biti "Oprema" ili "Dijelovi".' if strict else "",
        )
        ws.add_data_validation(dv)
        return dv

    cat_col = col_map.get("category", "C")   # slovo stupca za category

    if "category" in col_map:
        dv = make_dv("NR_Kategorija", "Kategorija",
                     "Odaberi: Oprema ili Dijelovi", strict=True)
        dv.sqref = f"{cat_col}2:{cat_col}{MAX_ROW}"

    if "subcategory" in col_map:
        # INDIRECT($C2) → kad je C2="Oprema" Excel uzima named range "Oprema"
        #                  kad je C2="Dijelovi" Excel uzima named range "Dijelovi"
        # Dropdown se automatski mijenja ovisno o odabranoj kategoriji!
        formula = f"INDIRECT(${cat_col}2)"
        dv = make_dv(formula, "Podkategorija",
                     "Podkategorije ovise o odabranoj kategoriji", strict=False)
        dv.sqref = f"{col_map['subcategory']}2:{col_map['subcategory']}{MAX_ROW}"

    if "brand" in col_map:
        # INDIRECT("NR_Brand_"&$C2) → "NR_Brand_Oprema" ili "NR_Brand_Dijelovi"
        formula = f'INDIRECT("NR_Brand_"&${cat_col}2)'
        dv = make_dv(formula, "Brand",
                     "Brendovi ovise o odabranoj kategoriji", strict=False)
        dv.sqref = f"{col_map['brand']}2:{col_map['brand']}{MAX_ROW}"

    active = [k for k in ("category", "subcategory", "brand") if k in col_map]
    print(f"  Padajući izbornici: {', '.join(active)}")

# ── AutoFilter ────────────────────────────────────────────────────────────────
def add_autofilter(ws):
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

# ── Brand sheet (Brendovi_Oprema / Brendovi_Dijelovi) ─────────────────────────
def write_brand_sheet(wb, sheet_name, brendovi, color):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    fill, font, center, border = header_style(color)
    db = data_border()

    for ci, col in enumerate(BRAND_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = font; cell.fill = fill
        cell.alignment = center; cell.border = border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    col_widths = {"brand_id":14,"name":26,"logo_folder":16,"website":28,"active":10}
    for ri, row_data in enumerate(brendovi, start=2):
        rf = FILLS[ri % 2]
        for ci, val in enumerate(row_data, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = rf; c.border = db
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(BRAND_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[col]

    add_autofilter(ws)
    print(f"  Sheet '{sheet_name}': {len(brendovi)} brendova")

# ── brand stupac u Proizvodi ──────────────────────────────────────────────────
def ensure_brand_column(ws):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if "brand" in headers:
        return
    next_col = ws.max_column + 1
    fill, font, center, border = header_style("1B5E20")
    cell = ws.cell(row=1, column=next_col, value="brand")
    cell.font = font; cell.fill = fill; cell.alignment = center; cell.border = border
    db = data_border()
    for ri in range(2, ws.max_row + 1):
        c = ws.cell(row=ri, column=next_col, value="")
        c.border = db; c.alignment = Alignment(vertical="center")
    ws.column_dimensions[get_column_letter(next_col)].width = 14
    print(f"  Stupac 'brand' dodan u Proizvodi (kolona {get_column_letter(next_col)})")

# ── Kreira Proizvodi sheet (učita xlsx ili regenerira iz products.json) ────────
PROD_HEADERS = ["id","name","category","subcategory","price",
                "description","image_folder","is_new","brand"]
PROD_COL_WIDTHS = {"id":5,"name":35,"category":12,"subcategory":18,
                   "price":16,"description":55,"image_folder":25,"is_new":8,"brand":14}

def _write_prod_headers(ws):
    fill, font, center, border = header_style("1B5E20")
    db = data_border()
    for ci, col in enumerate(PROD_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = font; cell.fill = fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = PROD_COL_WIDTHS.get(col, 15)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    return db

def _fill_from_json(ws, products):
    db = data_border()
    for row_idx, p in enumerate(products, start=2):
        rf = FILLS[row_idx % 2]
        row_data = {
            "id":           p.get("id", row_idx - 1),
            "name":         p.get("name", ""),
            "category":     p.get("category", ""),
            "subcategory":  p.get("subcategory", ""),
            "price":        p.get("price", ""),
            "description":  p.get("description", ""),
            "image_folder": p.get("image_folder", ""),
            "is_new":       p.get("is_new", False),
            "brand":        p.get("brand", ""),
        }
        for ci, col in enumerate(PROD_HEADERS, start=1):
            c = ws.cell(row=row_idx, column=ci, value=row_data[col])
            c.fill = rf; c.border = db
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 18

def ensure_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        print(f"Učitan: {EXCEL_PATH}")
    else:
        # xlsx ne postoji — kreiraj i popuni iz products.json
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proizvodi"
        _write_prod_headers(ws)

        products = []
        if os.path.exists(JSON_PATH):
            with open(JSON_PATH, encoding="utf-8") as f:
                products = json.load(f)

        if products:
            _fill_from_json(ws, products)
            print(f"Kreiran iz products.json: {EXCEL_PATH} ({len(products)} proizvoda)")
        else:
            print(f"Kreiran prazan: {EXCEL_PATH}  (products.json je prazan)")
    return wb

# ── Upute sheet ───────────────────────────────────────────────────────────────
def ensure_upute(wb):
    ws_u = wb["Upute"] if "Upute" in wb.sheetnames else wb.create_sheet("Upute")
    ws_u.delete_rows(1, ws_u.max_row)

    upute = [
        ("UPUTE ZA UPRAVLJANJE OPREMOM I DIJELOVIMA", ""),
        ("", ""),
        ("STUPCI (sheet Proizvodi)", ""),
        ("id",           "Redni broj — NE mijenjaj!"),
        ("name",         "Naziv proizvoda"),
        ("category",     "Padajući izbornik → Oprema  ili  Dijelovi"),
        ("subcategory",  "Padajući izbornik → ili slobodno upiši novu (npr. Kacige)"),
        ("price",        "Cijena npr: 149,90 €  ili  Cijena na upit"),
        ("description",  "Kratki opis (1-2 rečenice)"),
        ("image_folder", "Naziv mape — slike u public/images/oprema/ ili /dijelovi/"),
        ("is_new",       "TRUE = prikazuje badge NOVO,  FALSE = ne"),
        ("brand",        "Padajući izbornik → brand_id (npr. airoh, motul)"),
        ("", ""),
        ("DODAVANJE NOVE PODKATEGORIJE", ""),
        ("1.", f"Otvori sheet '{DD_SHEET}' → kolona B (Podkategorija)"),
        ("2.", "Dodaj novi naziv ispod zadnjeg reda"),
        ("3.", "Odmah se pojavljuje u padajućem izborniku — bez pokretanja skripte!"),
        ("", ""),
        ("DODAVANJE NOVOG BRENDA", ""),
        ("1.", f"Dodaj red u Brendovi_Oprema ili Brendovi_Dijelovi sheet"),
        ("2.", f"Dodaj isti brand_id u '{DD_SHEET}' → kolona C ili D"),
        ("3.", "Stavi logo u public/images/brendovi/{brand_id}.svg ili .png"),
        ("4.", "Pokreni: py -X utf8 scripts/kreiraj_excel_predlozak.py"),
        ("5.", "Pokreni: py -X utf8 scripts/proizvodi_excel_to_json.py"),
        ("", ""),
        ("DODAVANJE PROIZVODA", ""),
        ("1.", "Dodaj novi red u sheet Proizvodi"),
        ("2.", "Popuni stupce — category i brand odaberi iz padajućeg izbornika"),
        ("3.", "Stavi slike u public/images/oprema/{image_folder}/ ili /dijelovi/"),
        ("4.", "Pokreni: py -X utf8 scripts/proizvodi_excel_to_json.py"),
        ("5.", "git add . && git commit -m 'Dodan proizvod' && git push"),
        ("", ""),
        ("FILTER", "Strelice u zaglavlju → klikni za sortiranje i filtriranje unutar Excela"),
    ]

    for r, (a, b) in enumerate(upute, start=1):
        bold = bool(a and (a.isupper() or a.endswith(".")))
        fa = Font(bold=bold, size=9)
        if r == 1:
            fa = Font(bold=True, size=13, color="1B5E20")
        ws_u.cell(row=r, column=1, value=a).font = fa
        ws_u.cell(row=r, column=2, value=b).font = Font(size=9)

    ws_u.column_dimensions["A"].width = 36
    ws_u.column_dimensions["B"].width = 65

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = ensure_workbook()

    if "Proizvodi" in wb.sheetnames:
        ws_prod = wb["Proizvodi"]
        ensure_brand_column(ws_prod)
        add_autofilter(ws_prod)
    else:
        ws_prod = None
        print("  UPOZORENJE: Sheet 'Proizvodi' ne postoji")

    write_dropdown_sheet(wb)

    if ws_prod is not None:
        add_dropdowns(ws_prod)

    write_brand_sheet(wb, "Brendovi_Oprema",   BRENDOVI_OPREMA,   "E65100")
    write_brand_sheet(wb, "Brendovi_Dijelovi", BRENDOVI_DIJELOVI, "1565C0")
    ensure_upute(wb)

    # Spremi — direktno, fallback na tmp ako je fajl zaključan
    tmp = EXCEL_PATH + ".tmp"
    if os.path.exists(tmp):
        os.remove(tmp)
    try:
        wb.save(EXCEL_PATH)
        print(f"\nSačuvano: {EXCEL_PATH}")
    except PermissionError:
        wb.save(tmp)
        try:
            os.replace(tmp, EXCEL_PATH)
            print(f"\nSačuvano: {EXCEL_PATH}")
        except PermissionError:
            print(f"\nZATVORI Excel pa pokreni skriptu ponovo.")
            return

    print(f"\nOtvori Excel — sheet '{DD_SHEET}' ima sve padajuće liste.")
    print("Dodaj novi red u bilo koji stupac → automatski se pojavljuje u dropdownu.")


if __name__ == "__main__":
    main()
