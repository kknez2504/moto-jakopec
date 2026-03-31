"""
Kreira osnovni proizvodi.xlsx iz products.json (za inicijalizaciju / oporavak).

Nakon ovoga pokreni i predložak da dobiješ dropdowne:
    py -X utf8 scripts/proizvodi_json_to_excel.py
    py -X utf8 scripts/kreiraj_excel_predlozak.py

Stupci:
  id, name, category, subcategory, price, description, image_folder, is_new, brand
"""

import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR    = os.path.join(SCRIPT_DIR, "..")
JSON_PATH   = os.path.join(ROOT_DIR, "src", "data", "products.json")
EXCEL_PATH  = os.path.join(ROOT_DIR, "proizvodi.xlsx")

HEADERS = ["id", "name", "category", "subcategory", "price", "description", "image_folder", "is_new", "brand"]

COL_WIDTHS = {
    "id": 5, "name": 35, "category": 12, "subcategory": 18,
    "price": 16, "description": 55, "image_folder": 25, "is_new": 8, "brand": 14,
}

def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        products = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proizvodi"

    # ── Stil zaglavlja ─────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, col in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    fills = [
        PatternFill("solid", fgColor="F9FBF9"),
        PatternFill("solid", fgColor="E8F5E9"),
    ]

    for row_idx, p in enumerate(products, start=2):
        fill = fills[row_idx % 2]
        row_data = {
            "id":           p.get("id"),
            "name":         p.get("name", ""),
            "category":     p.get("category", ""),
            "subcategory":  p.get("subcategory", ""),
            "price":        p.get("price", ""),
            "description":  p.get("description", ""),
            "image_folder": p.get("image_folder", ""),
            "is_new":       p.get("is_new", False),
            "brand":        p.get("brand", ""),
        }
        for col_idx, col in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 18

    for col_idx, col in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 20)

    # ── Upute sheet ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Upute")
    ws2["A1"] = "UPUTE ZA UPRAVLJANJE OPREMOM I DIJELOVIMA"
    ws2["A1"].font = Font(bold=True, size=14, color="1B5E20")

    upute = [
        ("", ""),
        ("STUPCI", ""),
        ("id",           "Redni broj — ne mijenjaj!"),
        ("name",         "Naziv proizvoda (npr. Kawasaki kaciga KRT)"),
        ("category",     "Mora biti točno: Oprema   ili   Dijelovi"),
        ("subcategory",  "Podkategorija — slobodno piši (npr: Kacige / Jakne / Rukavice / Gume / Filteri / Ulja)"),
        ("price",        "Cijena npr: 149,90 € ili: Cijena na upit"),
        ("description",  "Kratki opis (1-2 rečenice)"),
        ("image_folder", "Naziv mape (npr. kaciga-krt) — slike idu u public/images/oprema/ ili public/images/dijelovi/"),
        ("is_new",       "TRUE = prikazuje se badge NOVO, FALSE = ne"),
        ("", ""),
        ("DODAVANJE PROIZVODA", ""),
        ("1.", "Dodaj novi red u Excel"),
        ("2.", "Ispuni sve stupce (category mora biti 'Oprema' ili 'Dijelovi')"),
        ("3.", "Napravi mapu u public/images/oprema/naziv-mape/ ili public/images/dijelovi/naziv-mape/ i stavi slike"),
        ("4.", "Pokreni: py -X utf8 scripts/proizvodi_excel_to_json.py"),
        ("5.", "git add . && git commit -m 'Dodan proizvod' && git push"),
        ("", ""),
        ("STRANICE NA WEBU", ""),
        ("Oprema",    "→ /oprema  (filtrira po category='Oprema')"),
        ("Dijelovi",  "→ /dijelovi  (filtrira po category='Dijelovi')"),
        ("", ""),
        ("NAPOMENA", "Slike iz public/images se automatski učitavaju — ne trebaš ništa pisati u JSON"),
    ]

    for r, (a, b) in enumerate(upute, start=2):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=bool(a and (a.isupper() or a.endswith("."))))
        ws2.cell(row=r, column=2, value=b)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 70

    wb.save(EXCEL_PATH)
    print(f"Spremljeno: {EXCEL_PATH}")
    print(f"  Proizvoda: {len(products)}")


if __name__ == "__main__":
    main()
