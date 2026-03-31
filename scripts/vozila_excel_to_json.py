"""
Pretvara vozila.xlsx natrag u models.json.

Pokreni nakon svake izmjene u Excelu:
    py -X utf8 scripts/vozila_excel_to_json.py

Što radi:
  - Čita vozila.xlsx (sheet "Vozila")
  - Fiksni stupci → osnovni podaci vozila
  - Ostali stupci → specs (prazne ćelije se preskaču)
  - image_folder → automatski učitava slike iz public/images/vozila/{folder}/
  - Sprema src/data/models.json
"""

import json, os, re
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.join(SCRIPT_DIR, "..")
EXCEL_PATH = os.path.join(ROOT_DIR, "vozila.xlsx")
JSON_PATH  = os.path.join(ROOT_DIR, "src", "data", "models.json")
IMG_DIR    = os.path.join(ROOT_DIR, "public", "images", "vozila")

FIXED_COLS = {"id", "name", "category", "license_category",
              "is_new", "price", "description", "image_folder"}

# ── Sortiranje slika ──────────────────────────────────────────────────────────
def img_sort_key(fname: str) -> str:
    """
    Prioritet:
      _STU_ / bocna prednji kraj desno  → 0_
      _STV_ / bocna prednji kraj lijevo → 2_
      _ACT_                             → 3_
      FT (detail shot motorcycles)      → 7_
      _DET_                             → 8_
      sve ostalo                        → 5_
    """
    u = fname.upper()
    if "_STU_" in u:  return "0_" + fname
    if "_STV_" in u:  return "2_" + fname
    if "_ACT_" in u:  return "3_" + fname
    if re.search(r'FT\d*\.(jpg|png|webp)$', u): return "7_" + fname
    if "_DET_" in u:  return "8_" + fname
    return "5_" + fname

def get_images(image_folder: str) -> list:
    if not image_folder:
        return []
    folder_path = os.path.join(IMG_DIR, image_folder)
    if not os.path.isdir(folder_path):
        return []
    exts = {".jpg", ".jpeg", ".png", ".webp"}
    files = [f for f in os.listdir(folder_path)
             if os.path.splitext(f)[1].lower() in exts]
    files.sort(key=img_sort_key)
    return [f"/images/vozila/{image_folder}/{f}" for f in files]

# ── Parsiranje vrijednosti ─────────────────────────────────────────────────────
def parse_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() in ("TRUE", "DA", "1", "YES")
    if isinstance(val, (int, float)):
        return bool(val)
    return False

def parse_value(val):
    """Vraća čistu vrijednost ili None ako je prazno."""
    if val is None:
        return None
    if isinstance(val, str):
        stripped = val.strip()
        return stripped if stripped else None
    return val

# ── Glavni dio ────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"GREŠKA: Datoteka nije pronađena: {EXCEL_PATH}")
        print("Pokreni prvo: py -X utf8 scripts/vozila_json_to_excel.py")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    if "Vozila" not in wb.sheetnames:
        print("GREŠKA: Sheet 'Vozila' ne postoji u Excel datoteci!")
        return

    ws = wb["Vozila"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("GREŠKA: Excel je prazan!")
        return

    # Zaglavlje
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    spec_cols = [h for h in headers if h and h not in FIXED_COLS]

    print(f"Stupci: {len(headers)} ({len(spec_cols)} spec stupaca)")

    models = []
    skipped = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}

        # Preskoči potpuno prazne redove
        name = parse_value(row_dict.get("name"))
        if not name:
            skipped += 1
            continue

        # Osnovni podaci
        raw_id = row_dict.get("id")
        try:
            model_id = int(raw_id) if raw_id is not None else row_idx - 1
        except (ValueError, TypeError):
            model_id = row_idx - 1

        image_folder = parse_value(row_dict.get("image_folder")) or ""
        images = get_images(image_folder)

        model = {
            "id":               model_id,
            "name":             name,
            "category":         parse_value(row_dict.get("category")) or "",
            "license_category": parse_value(row_dict.get("license_category")) or "",
            "is_new":           parse_bool(row_dict.get("is_new", False)),
            "price":            parse_value(row_dict.get("price")) or "",
            "description":      parse_value(row_dict.get("description")) or "",
            "images":           images,
        }

        # Specs — samo neprazna polja
        specs = {}
        for key in spec_cols:
            val = parse_value(row_dict.get(key))
            if val is not None:
                specs[key] = str(val)
        if specs:
            model["specs"] = specs

        models.append(model)

    # Spremi JSON
    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(models, f, ensure_ascii=False, indent=2)

    print(f"Spremljeno: {JSON_PATH}")
    print(f"  Vozila: {len(models)}")
    if skipped:
        print(f"  Preskočeni prazni redovi: {skipped}")
    print()
    print("Sljedeći koraci:")
    print("  git add src/data/models.json")
    print("  git commit -m 'Ažurirani podaci vozila'")
    print("  git push")

if __name__ == "__main__":
    main()
