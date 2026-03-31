"""
Pretvara models.json u vozila.xlsx (jednom za inicijalizaciju).

Pokreni: py -X utf8 scripts/json_to_excel.py
"""

import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.join(SCRIPT_DIR, "..")
JSON_PATH  = os.path.join(ROOT_DIR, "src", "data", "models.json")
EXCEL_PATH = os.path.join(ROOT_DIR, "vozila.xlsx")
IMG_DIR    = os.path.join(ROOT_DIR, "public", "images", "vozila")

# Fiksni stupci (uvijek prisutni)
FIXED_COLS = ["id", "name", "category", "license_category", "is_new", "price", "description", "image_folder"]

def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        models = json.load(f)

    # Skupi sve spec kljuceve (zadrzava redoslijed pojavljivanja)
    all_spec_keys = []
    seen = set()
    for m in models:
        for k in m.get("specs", {}).keys():
            if k not in seen:
                all_spec_keys.append(k)
                seen.add(k)

    headers = FIXED_COLS + all_spec_keys
    total_cols = len(headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vozila"

    # ── Stil zaglavlja ────────────────────────────────────────────────
    header_fill_main  = PatternFill("solid", fgColor="1B5E20")   # tamno zelena - fiksni stupci
    header_fill_specs = PatternFill("solid", fgColor="2E7D32")   # srednje zelena - spec stupci
    header_font  = Font(color="FFFFFF", bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Zagljavlje
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill_main if col_idx <= len(FIXED_COLS) else header_fill_specs
        cell.alignment = center_align
        cell.border = thin_border

    ws.freeze_panes = "A2"  # zamrzni zaglavlje
    ws.row_dimensions[1].height = 30

    # ── Boje alternativnih redova ─────────────────────────────────────
    fills = [
        PatternFill("solid", fgColor="F9FBF9"),
        PatternFill("solid", fgColor="E8F5E9"),
    ]

    # ── Podaci ───────────────────────────────────────────────────────
    for row_idx, m in enumerate(models, start=2):
        fill = fills[row_idx % 2]

        # Pronadi mapu slika za ovaj model
        image_folder = ""
        images = m.get("images", [])
        if images:
            # Npr. "/images/01-ninja-h2r/img_01.jpg" → "01-ninja-h2r"
            parts = images[0].split("/")
            if len(parts) >= 3:
                image_folder = parts[2]

        row_data = {
            "id":               m.get("id"),
            "name":             m.get("name", ""),
            "category":         m.get("category", ""),
            "license_category": m.get("license_category", ""),
            "is_new":           m.get("is_new", False),
            "price":            m.get("price", ""),
            "description":      m.get("description", ""),
            "image_folder":     image_folder,
        }
        # Dodaj spec vrijednosti
        for k in all_spec_keys:
            row_data[k] = m.get("specs", {}).get(k, "")

        for col_idx, col_name in enumerate(headers, start=1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        ws.row_dimensions[row_idx].height = 18

    # ── Širina stupaca ────────────────────────────────────────────────
    col_widths = {
        "id": 5, "name": 30, "category": 15, "license_category": 8,
        "is_new": 8, "price": 14, "description": 50, "image_folder": 22,
    }
    for col_idx, col_name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        w = col_widths.get(col_name, 22)
        ws.column_dimensions[letter].width = w

    # ── Fiksni stupci vizualno odvojeni ──────────────────────────────
    # (nic posebnog nije potrebno, boja zaglavlja vec dijeli)

    # ── Legenda u Sheet2 ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Upute")
    ws2["A1"] = "UPUTE ZA KORIŠTENJE"
    ws2["A1"].font = Font(bold=True, size=14, color="1B5E20")
    upute = [
        ("", ""),
        ("ZELENI STUPCI", "Osnovni podaci vozila — uredi slobodno"),
        ("id",            "Redni broj (ne mijenjaj redoslijed!)"),
        ("name",          "Puno ime vozila (npr. Kawasaki Z 900)"),
        ("category",      "Kategorija: Sport / Naked / Adventure / Touring / Classic / Cruiser / Offroad / Električni / Quad / Mule / Jet Ski"),
        ("license_category", "Kategorija vozačke (A1, A2 ili prazno za neograničenu)"),
        ("is_new",        "TRUE = novo, FALSE = nije novo (prikazuje se badge NOVO)"),
        ("price",         "Cijena npr: 9.462 € ili: Cijena na upit"),
        ("description",   "Kratki opis vozila (1-2 rečenice)"),
        ("image_folder",  "Naziv mape u public/images/vozila/ (npr. 05-ninja-zx-4rr)"),
        ("", ""),
        ("TAMNOZELENI STUPCI", "Tehničke specifikacije — uredi ili ostavi prazno"),
        ("",              "Prazna polja = NE prikazuju se na stranici"),
        ("", ""),
        ("DODAVANJE NOVOG VOZILA", ""),
        ("1.", "Dodaj novi red ispod zadnjeg vozila"),
        ("2.", "Ispuni sve zelene stupce"),
        ("3.", "Stavi slike u public/images/vozila/XX-naziv-vozila/"),
        ("4.", "Upiši naziv te mape u stupac image_folder"),
        ("5.", "Pokreni: py -X utf8 scripts/vozila_excel_to_json.py"),
        ("6.", "git add . && git commit -m 'Dodano vozilo' && git push"),
        ("", ""),
        ("BRISANJE VOZILA", "Obriši cijeli red — pa pokreni skriptu i pusaj"),
        ("PROMJENA CIJENE",  "Promijeni vrijednost u stupcu price — pa pokreni skriptu i pusaj"),
    ]
    for r, (a, b) in enumerate(upute, start=2):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=bool(a and a.isupper()))
        ws2.cell(row=r, column=2, value=b)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 70

    wb.save(EXCEL_PATH)
    print(f"Saved: {EXCEL_PATH}")
    print(f"  Vozila: {len(models)}")
    print(f"  Spec stupci: {len(all_spec_keys)}")
    print(f"  Ukupno stupaca: {total_cols}")

if __name__ == "__main__":
    main()
